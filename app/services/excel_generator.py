from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


HEADER_FILL = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='ffffff', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, color='1a237e', size=14)
NORMAL_FONT = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)
BALANCE_POSITIVE = Font(name='Calibri', bold=True, color='1b5e20', size=10)
BALANCE_NEGATIVE = Font(name='Calibri', bold=True, color='b71c1c', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='middle')
LEFT_ALIGN = Alignment(horizontal='left', vertical='middle')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='middle')


def generate_ledger_excel(entries, party_name='', business_name='My Business',
                          opening_balance=0, closing_balance=0,
                          date_from='', date_to=''):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ledger Statement'

    ws.merge_cells('A1:F1')
    ws['A1'] = business_name.upper()
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='middle')

    ws.merge_cells('A2:F2')
    title = 'Ledger Statement'
    if party_name:
        title += f' - {party_name}'
    ws['A2'] = title
    ws['A2'].font = Font(name='Calibri', bold=True, color='333333', size=12)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='middle')

    if date_from or date_to:
        ws.merge_cells('A3:F3')
        date_range = f'Period: {date_from} to {date_to}' if date_from and date_to else f'From: {date_from or "Start"} To: {date_to or "End"}'
        ws['A3'] = date_range
        ws['A3'].font = Font(name='Calibri', color='666666', size=9)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='middle')

    start_row = 5
    headers = ['Date', 'Type', 'Ref No', 'Debit', 'Credit', 'Balance']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    row = start_row + 1
    cell = ws.cell(row=row, column=1, value='Opening Balance')
    cell.font = BOLD_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = THIN_BORDER
    cell = ws.cell(row=row, column=6, value=opening_balance)
    cell.font = BOLD_FONT
    cell.number_format = '#,##0.00'
    cell.alignment = RIGHT_ALIGN
    cell.border = THIN_BORDER

    for entry in entries:
        row += 1
        date_val = entry['date'].strftime('%d-%m-%Y') if hasattr(entry['date'], 'strftime') else str(entry['date'])
        ws.cell(row=row, column=1, value=date_val).font = NORMAL_FONT
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        type_font = Font(name='Calibri', size=10, bold=True,
                         color='1565c0' if entry['type'] == 'Invoice' else 'e65100')
        ws.cell(row=row, column=2, value=entry['type']).font = type_font
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2).border = THIN_BORDER

        ws.cell(row=row, column=3, value=entry['ref_number']).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).border = THIN_BORDER

        if entry['debit'] > 0:
            cell = ws.cell(row=row, column=4, value=entry['debit'])
            cell.font = NORMAL_FONT
            cell.number_format = '#,##0.00'
            cell.alignment = RIGHT_ALIGN
        ws.cell(row=row, column=4).border = THIN_BORDER

        if entry['credit'] > 0:
            cell = ws.cell(row=row, column=5, value=entry['credit'])
            cell.font = NORMAL_FONT
            cell.number_format = '#,##0.00'
            cell.alignment = RIGHT_ALIGN
        ws.cell(row=row, column=5).border = THIN_BORDER

        cell = ws.cell(row=row, column=6, value=entry['balance'])
        cell.font = BALANCE_POSITIVE if entry['balance'] >= 0 else BALANCE_NEGATIVE
        cell.number_format = '#,##0.00'
        cell.alignment = RIGHT_ALIGN
        cell.border = THIN_BORDER

    row += 1
    ws.cell(row=row, column=1, value='Closing Balance').font = BOLD_FONT
    ws.cell(row=row, column=1).alignment = LEFT_ALIGN
    ws.cell(row=row, column=1).border = Border(
        top=Side(style='double', color='1a237e'),
        bottom=Side(style='double', color='1a237e'),
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc')
    )
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = Border(
            top=Side(style='double', color='1a237e'),
            bottom=Side(style='double', color='1a237e')
        )
    ws.cell(row=row, column=5).border = Border(
        top=Side(style='double', color='1a237e'),
        bottom=Side(style='double', color='1a237e')
    )
    cell = ws.cell(row=row, column=6, value=closing_balance)
    cell.font = Font(name='Calibri', bold=True, color='1a237e', size=11)
    cell.number_format = '#,##0.00'
    cell.alignment = RIGHT_ALIGN
    cell.border = Border(
        top=Side(style='double', color='1a237e'),
        bottom=Side(style='double', color='1a237e'),
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc')
    )

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
